from PyPDF2 import PdfFileReader
from docx import Document
from openpyxl import load_workbook

from PyPDF2 import PdfReader


def extract_pdf_preview(uploaded_file):
    # Placeholder function for extracting a preview from a PDF file
    # Replace this with actual implementation using a library like PyPDF2

    # For demonstration, we'll simply return the first few lines of text from the PDF
    try:
        pdf_reader = PdfReader(uploaded_file)
        num_pages = min(len(pdf_reader.pages), 3)  # Extract text from the first 3 pages
        preview_lines = []
        for page_num in range(num_pages):
            preview_lines.append(pdf_reader.pages[page_num].extract_text())
        return '\n'.join(preview_lines)
    except Exception as e:
        print("Error reading PDF:", e)
        return ""


def extract_docx_preview(uploaded_file):
    # Placeholder function for extracting a preview from a DOCX file
    # Replace this with actual implementation using the python-docx library

    # For demonstration, we'll simply return the first few paragraphs of text from the DOCX
    try:
        doc = Document(uploaded_file)
        preview_paragraphs = [paragraph.text for paragraph in doc.paragraphs[:3]]  # Extract first 3 paragraphs
        return '\n'.join(preview_paragraphs)
    except Exception as e:
        print("Error reading DOCX:", e)
        return ""


def extract_xlsx_preview(uploaded_file):
    # Placeholder function for extracting a preview from an XLSX file
    # Replace this with actual implementation using a library like openpyxl

    # For demonstration, we'll simply return the values from the first few cells of the first sheet
    try:
        wb = load_workbook(uploaded_file)
        ws = wb.active
        preview_values = []
        for row in ws.iter_rows(max_row=3):  # Extract values from the first 3 rows
            row_values = [str(cell.value) for cell in row]
            preview_values.append(' '.join(row_values))
        return '\n'.join(preview_values)
    except Exception as e:
        print("Error reading XLSX:", e)
        return ""
